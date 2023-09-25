import openpyxl

#	.xlsx file -->	Workbook->	sheet->	 rows& coloumn---> cells

xlpath = r"C:\Users\user\Desktop\sample.xlsx"
wb=openpyxl.load_workbook(xlpath)
sheet=wb["Sheet3"]

print("number of rows", n_row)
print("number of column", n_col)

sheet.cell(1,1).value="emp1"
sheet.cell(1,2).value="123"

sheet.cell(2,1).value="emp2"
sheet.cell(2,2).value="345"

sheet.cell(3,1).value="emp3"
sheet.cell(3,2).value="567"

sheet.cell(4,1).value="emp4"
sheet.cell(4,2).value="789"

wb.save(xlpath)